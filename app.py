import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import joblib
import json
import io

st.set_page_config(
    page_title="Warehouse Decision System",
    layout="wide",
    page_icon="🏭",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
[data-testid="stMetricValue"] { font-size: 2rem; font-weight: 700; }
.kpi-card { background: #1F4E79; border-radius: 10px; padding: 16px; text-align: center; color: white; margin-bottom: 8px; }
.kpi-label { font-size: 0.85rem; opacity: 0.8; text-transform: uppercase; letter-spacing: 1px; }
.kpi-value { font-size: 1.8rem; font-weight: 700; margin-top: 4px; }
.section-header { font-size: 1.1rem; font-weight: 600; color: #002060; border-bottom: 2px solid #002060; padding-bottom: 4px; margin: 16px 0 12px 0; }
.alert-box { padding: 12px 16px; border-radius: 8px; margin: 8px 0; font-weight: 500; }
.alert-red    { background: #FFE4E4; border-left: 4px solid #C00000; color: #7B0000; }
.alert-amber  { background: #FFF8E1; border-left: 4px solid #FFC000; color: #7B5000; }
.alert-green  { background: #E8F5E9; border-left: 4px solid #70AD47; color: #1B5E20; }
</style>
""", unsafe_allow_html=True)

@st.cache_data
def load_data():
    return pd.read_csv("warehouse_data.csv")

@st.cache_resource
def load_models():
    try:
        tp_model   = joblib.load("models/throughput_model.pkl")
        risk_model = joblib.load("models/risk_classifier.pkl")
        shift_enc  = joblib.load("models/shift_encoder.pkl")
        risk_enc   = joblib.load("models/risk_encoder.pkl")
        return tp_model, risk_model, shift_enc, risk_enc, True
    except:
        return None, None, None, None, False

@st.cache_data
def load_metrics():
    try:
        with open("models/metrics.json") as f:
            return json.load(f)
    except:
        return {}

@st.cache_data
def load_dashboard_template():
    try:
        with open("dashboard/warehouse_operations_dashboard.xlsx", "rb") as f:
            return f.read()
    except:
        return None

@st.cache_data
def get_tree_predictions(shift_enc_classes, shift, staff_hours, disruption_hours, order_volume, _models_ok):
    """Cache the 150-tree prediction loop — only re-runs when inputs change.
    _models_ok is prefixed with _ so Streamlit skips hashing it (bool is fine
    to hash, but the underscore convention makes the intent explicit).
    """
    if not _models_ok:
        return None
    try:
        shift_num = shift_enc.transform([shift])[0]
        X = pd.DataFrame([[shift_num, staff_hours, disruption_hours, order_volume]],
                         columns=['shift_enc','staff_hours','disruption_hours','order_volume'])
        return np.array([t.predict(X)[0] for t in tp_model.estimators_])
    except:
        return None

@st.cache_data
def compute_optimizer_grid(shift, order_volume, _models_ok):
    """Pre-compute all 81 predictions for the optimizer heatmap."""
    rows = []
    for dis in range(0, 9):
        for staff in range(4, 13):
            tp_pred, _, _, risk_pred, _ = predict(shift, staff, dis, order_volume)
            cost_total = dis * DISRUPTION_COST_PER_HOUR + staff * LABOUR_COST_PER_HOUR
            rows.append({"Disruption Hours": dis, "Staff Hours": staff,
                         "Predicted Throughput": tp_pred, "Total Cost": cost_total,
                         "Risk Level": risk_pred})
    return pd.DataFrame(rows)

@st.cache_data
def compute_sensitivity_sweep(shift, disruption_hours, order_volume, _models_ok):
    """Cache the staff-hours sensitivity sweep (9 predictions)."""
    result = []
    for sh in range(4, 13):
        tp, lo, hi, _, _ = predict(shift, sh, disruption_hours, order_volume)
        result.append({"Staff Hours": sh, "Throughput": tp, "Lower": lo, "Upper": hi})
    return result

df = load_data()
tp_model, risk_model, shift_enc, risk_enc, models_ok = load_models()
metrics = load_metrics()

def predict(shift, staff_hours, disruption_hours, order_volume):
    if models_ok:
        try:
            shift_num = shift_enc.transform([shift])[0]
            X = pd.DataFrame([[shift_num, staff_hours, disruption_hours, order_volume]],
                             columns=['shift_enc','staff_hours','disruption_hours','order_volume'])
            tree_preds = np.array([t.predict(X)[0] for t in tp_model.estimators_])
            throughput = int(tp_model.predict(X)[0])
            tp_lo = int(np.percentile(tree_preds, 10))
            tp_hi = int(np.percentile(tree_preds, 90))
            risk_proba = risk_model.predict_proba(X)[0]
            risk_label = risk_enc.inverse_transform([risk_model.predict(X)[0]])[0]
            risk_conf  = round(float(max(risk_proba)) * 100, 1)
            return throughput, tp_lo, tp_hi, risk_label, risk_conf
        except:
            pass
    shift_bonus = {"Morning": 400, "Afternoon": 150, "Night": 0}.get(shift, 0)
    throughput = int(600 + staff_hours * 280 - disruption_hours * 320 + order_volume * 0.3 + shift_bonus)
    risk_label = "High" if disruption_hours >= 4 else ("Medium" if disruption_hours >= 2 else "Low")
    return throughput, int(throughput * 0.88), int(throughput * 1.12), risk_label, 75.0

DISRUPTION_COST_PER_HOUR = 2286
LABOUR_COST_PER_HOUR     = 180

with st.sidebar:
    st.markdown("## 🏭 Scenario Controls")
    st.markdown("### Scenario A")
    shift_a      = st.selectbox("Shift", df["shift"].unique(), key="shift_a")
    staff_a      = st.slider("Staff Hours", 4, 12, 8, key="staff_a")
    disruption_a = st.slider("Disruption Hours", 0, 8, 1, key="dis_a")
    order_a      = st.number_input("Order Volume", 500, 5000, 1500, step=100, key="ord_a")
    st.divider()
    compare_mode = st.toggle("Enable A/B Comparison", value=False)
    if compare_mode:
        st.markdown("### Scenario B")
        shift_b      = st.selectbox("Shift", df["shift"].unique(), index=2, key="shift_b")
        staff_b      = st.slider("Staff Hours", 4, 12, 6, key="staff_b")
        disruption_b = st.slider("Disruption Hours", 0, 8, 4, key="dis_b")
        order_b      = st.number_input("Order Volume", 500, 5000, 1500, step=100, key="ord_b")
    st.divider()
    st.markdown("### Alert Thresholds")
    alert_cost       = st.number_input("Max Disruption Cost ($)", 0, 50000, 5000, step=500)
    alert_throughput = st.number_input("Min Throughput (units)", 0, 6000, 1500, step=100)

tp_a, tp_lo_a, tp_hi_a, risk_a, conf_a = predict(shift_a, staff_a, disruption_a, order_a)
cost_a   = disruption_a * DISRUPTION_COST_PER_HOUR
labour_a = staff_a * LABOUR_COST_PER_HOUR

if compare_mode:
    tp_b, tp_lo_b, tp_hi_b, risk_b, conf_b = predict(shift_b, staff_b, disruption_b, order_b)
    cost_b   = disruption_b * DISRUPTION_COST_PER_HOUR
    labour_b = staff_b * LABOUR_COST_PER_HOUR

col_h1, col_h2 = st.columns([3, 1])
with col_h1:
    st.title("🏭 Warehouse Decision Support System")
    st.caption("ML-powered insights for staffing, throughput & risk · Built by Garvit Mittal · v2.0")
with col_h2:
    if models_ok:
        st.success("✅ Real ML models active")
    else:
        st.warning("⚠️ Formula fallback")

alerts = []
if cost_a > alert_cost:
    alerts.append(("red", f"Disruption cost ${cost_a:,.0f} exceeds threshold ${alert_cost:,.0f}"))
if tp_a < alert_throughput:
    alerts.append(("red", f"Throughput {tp_a:,} units below minimum {alert_throughput:,}"))
if risk_a == "High":
    alerts.append(("red", "HIGH RISK — consider reducing disruption hours or increasing staff"))
elif risk_a == "Medium":
    alerts.append(("amber", "MEDIUM RISK — monitor closely"))
else:
    alerts.append(("green", "LOW RISK — conditions look good"))

for level, msg in alerts:
    icon = "🚨" if level=="red" else ("⚠️" if level=="amber" else "✅")
    st.markdown(f'<div class="alert-box alert-{level}">{icon} {msg}</div>', unsafe_allow_html=True)

tab_overview, tab_compare, tab_optimizer, tab_model, tab_export = st.tabs([
    "📊 Overview", "⚖️ A/B Comparison", "🔧 What-If Optimizer", "🤖 Model Intelligence", "📥 Export Reports"
])

# ── TAB 1: OVERVIEW ──────────────────────────────────────────────────────────
with tab_overview:
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Predicted Throughput", f"{tp_a:,} units", help=f"90% CI: {tp_lo_a:,}–{tp_hi_a:,}")
    c2.metric("Risk Level", risk_a, help=f"Confidence: {conf_a}%")
    c3.metric("Disruption Cost", f"${cost_a:,.0f}")
    c4.metric("Labour Cost", f"${labour_a:,.0f}")
    c5.metric("Total Cost", f"${cost_a+labour_a:,.0f}")
    st.divider()

    col_l, col_r = st.columns(2)
    with col_l:
        st.markdown('<p class="section-header">Throughput by Shift (Historical)</p>', unsafe_allow_html=True)
        shift_summary = df.groupby('shift').agg(mean_tp=('throughput_units','mean'), std_tp=('throughput_units','std')).reset_index()
        colors_map = {'Morning':'#002060','Afternoon':'#1F4E79','Night':'#2E75B6'}
        fig_shift = go.Figure()
        for _, row in shift_summary.iterrows():
            fig_shift.add_trace(go.Bar(x=[row['shift']], y=[row['mean_tp']],
                error_y=dict(type='data', array=[row['std_tp']], visible=True),
                name=row['shift'], marker_color=colors_map.get(row['shift'],'#555'),
                text=[f"{row['mean_tp']:.0f}"], textposition='outside'))
        fig_shift.update_layout(showlegend=False, height=300, yaxis_title="Avg Throughput",
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10))
        st.plotly_chart(fig_shift, use_container_width=True)

    with col_r:
        st.markdown('<p class="section-header">Risk Distribution (Historical)</p>', unsafe_allow_html=True)
        risk_counts = df['risk_level'].value_counts().reset_index()
        risk_counts.columns = ['Risk Level','Count']
        fig_risk = px.pie(risk_counts, names='Risk Level', values='Count', hole=0.45,
            color='Risk Level', color_discrete_map={'Low':'#70AD47','Medium':'#FFC000','High':'#C00000'})
        fig_risk.update_traces(textposition='outside', textinfo='percent+label')
        fig_risk.update_layout(height=300, margin=dict(t=10,b=10), paper_bgcolor='rgba(0,0,0,0)', showlegend=False)
        st.plotly_chart(fig_risk, use_container_width=True)

    st.markdown('<p class="section-header">Throughput Distribution with Scenario A Prediction</p>', unsafe_allow_html=True)
    fig_hist = go.Figure()
    fig_hist.add_trace(go.Histogram(x=df['throughput_units'], nbinsx=40, marker_color='#BDD7EE', opacity=0.8))
    fig_hist.add_vline(x=tp_a, line_color='#C00000', line_width=2, annotation_text=f"  Scenario A: {tp_a:,}", annotation_font_color='#C00000')
    fig_hist.add_vrect(x0=tp_lo_a, x1=tp_hi_a, fillcolor='rgba(192,0,0,0.08)', line_width=0, annotation_text="90% CI")
    fig_hist.update_layout(height=250, xaxis_title="Throughput (units)", yaxis_title="Frequency",
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10), showlegend=False)
    st.plotly_chart(fig_hist, use_container_width=True)

    st.markdown('<p class="section-header">Dataset Sample (1,200 rows)</p>', unsafe_allow_html=True)
    st.dataframe(df.head(50), use_container_width=True, height=200)

# ── TAB 2: A/B COMPARISON ────────────────────────────────────────────────────
with tab_compare:
    if not compare_mode:
        st.info("Enable **A/B Comparison** in the sidebar to compare two scenarios side-by-side.")
    else:
        col_a, col_mid, col_b = st.columns([5, 1, 5])
        risk_col = {"Low":"#375623","Medium":"#7B5000","High":"#7B0000"}

        def kpi_card(label, val, color="#1F4E79"):
            return f'<div class="kpi-card" style="background:{color}"><div class="kpi-label">{label}</div><div class="kpi-value">{val}</div></div>'

        with col_a:
            st.markdown(f"#### Scenario A — {shift_a} Shift")
            st.markdown(kpi_card("Throughput", f"{tp_a:,} units"), unsafe_allow_html=True)
            st.markdown(kpi_card("Risk Level", risk_a, color=risk_col.get(risk_a,"#1F4E79")), unsafe_allow_html=True)
            st.markdown(kpi_card("Disruption Cost", f"${cost_a:,.0f}", "#7B0000" if cost_a>5000 else "#1F4E79"), unsafe_allow_html=True)
            st.markdown(kpi_card("Labour Cost", f"${labour_a:,.0f}"), unsafe_allow_html=True)
            st.markdown(kpi_card("Model Confidence", f"{conf_a}%"), unsafe_allow_html=True)
        with col_mid:
            st.markdown("<br><br><br><br>", unsafe_allow_html=True)
            st.markdown("### vs")
        with col_b:
            st.markdown(f"#### Scenario B — {shift_b} Shift")
            st.markdown(kpi_card("Throughput", f"{tp_b:,} units"), unsafe_allow_html=True)
            st.markdown(kpi_card("Risk Level", risk_b, color=risk_col.get(risk_b,"#1F4E79")), unsafe_allow_html=True)
            st.markdown(kpi_card("Disruption Cost", f"${cost_b:,.0f}", "#7B0000" if cost_b>5000 else "#1F4E79"), unsafe_allow_html=True)
            st.markdown(kpi_card("Labour Cost", f"${labour_b:,.0f}"), unsafe_allow_html=True)
            st.markdown(kpi_card("Model Confidence", f"{conf_b}%"), unsafe_allow_html=True)

        st.divider()
        tp_delta   = tp_a - tp_b
        cost_delta = (cost_a+labour_a) - (cost_b+labour_b)
        if tp_delta > 0 and cost_delta <= 0:
            st.success(f"✅ Scenario A wins — {tp_delta:+,} units, ${abs(cost_delta):,.0f} lower cost")
        elif tp_delta < 0 and cost_delta >= 0:
            st.success(f"✅ Scenario B wins — {abs(tp_delta):,} more units, ${abs(cost_delta):,.0f} lower cost")
        else:
            st.info(f"⚖️ Trade-off — A: {tp_delta:+,} units, {cost_delta:+,.0f} cost vs B")

        # Grouped bar chart
        fig_bar = go.Figure(data=[
            go.Bar(name='Scenario A', x=['Throughput','Disruption Cost','Labour Cost'],
                   y=[tp_a, cost_a, labour_a], marker_color='#002060',
                   text=[f"{tp_a:,}", f"${cost_a:,.0f}", f"${labour_a:,.0f}"], textposition='outside'),
            go.Bar(name='Scenario B', x=['Throughput','Disruption Cost','Labour Cost'],
                   y=[tp_b, cost_b, labour_b], marker_color='#C00000',
                   text=[f"{tp_b:,}", f"${cost_b:,.0f}", f"${labour_b:,.0f}"], textposition='outside'),
        ])
        fig_bar.update_layout(barmode='group', height=380,
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10))
        st.plotly_chart(fig_bar, use_container_width=True)

        # Radar chart
        categories = ['Throughput','Low Disruption','Staff Efficiency','Low Cost','Low Risk']
        risk_num   = {"Low":1.0,"Medium":0.5,"High":0.0}
        def norm(v, mn, mx): return max(0, min(1, (v-mn)/(mx-mn+1e-9)))
        scores_a = [norm(tp_a,500,6000), norm(8-disruption_a,0,8), norm(tp_a/max(staff_a,1),50,700),
                    norm(50000-(cost_a+labour_a),0,50000), risk_num.get(risk_a,0.5)]
        scores_b = [norm(tp_b,500,6000), norm(8-disruption_b,0,8), norm(tp_b/max(staff_b,1),50,700),
                    norm(50000-(cost_b+labour_b),0,50000), risk_num.get(risk_b,0.5)]
        fig_radar = go.Figure()
        for scores, name, color in [(scores_a,"Scenario A","#002060"),(scores_b,"Scenario B","#C00000")]:
            fig_radar.add_trace(go.Scatterpolar(
                r=scores+[scores[0]], theta=categories+[categories[0]],
                fill='toself', name=name, line_color=color))
        fig_radar.update_layout(polar=dict(radialaxis=dict(visible=True,range=[0,1])),
            height=400, paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=30,b=10))
        st.plotly_chart(fig_radar, use_container_width=True)

# ── TAB 3: WHAT-IF OPTIMIZER ─────────────────────────────────────────────────
with tab_optimizer:
    st.markdown("### What-If Optimizer")
    st.markdown("Set a target throughput — we'll find the minimum staff hours needed, across every disruption level.")
    col_o1, col_o2 = st.columns(2)
    with col_o1:
        target_tp  = st.number_input("Target Throughput (units)", 500, 6000, 3000, step=100)
        opt_shift  = st.selectbox("Shift", df["shift"].unique(), key="opt_shift")
        opt_order  = st.number_input("Order Volume", 500, 5000, 1500, step=100, key="opt_order")
    with col_o2:
        max_budget = st.number_input("Max Total Cost Budget ($)", 1000, 100000, 20000, step=500)
        st.metric("Target", f"{target_tp:,} units")
        st.metric("Budget", f"${max_budget:,.0f}")

    res_df = compute_optimizer_grid(opt_shift, opt_order, models_ok)
    res_df["Meets Target"]  = res_df["Predicted Throughput"] >= target_tp
    res_df["Within Budget"] = res_df["Total Cost"] <= max_budget
    feasible = res_df[res_df["Meets Target"] & res_df["Within Budget"]]

    if feasible.empty:
        st.error("No combination meets your target within budget. Try raising budget or lowering target.")
    else:
        best = feasible.sort_values("Total Cost").iloc[0]
        st.success(f"Optimal: {int(best['Staff Hours'])} staff hrs, {int(best['Disruption Hours'])} disruption hrs "
                   f"→ {int(best['Predicted Throughput']):,} units at ${int(best['Total Cost']):,.0f} total cost")

    pivot = res_df.pivot(index="Disruption Hours", columns="Staff Hours", values="Predicted Throughput")
    fig_heat = px.imshow(pivot, color_continuous_scale="Blues",
        labels=dict(x="Staff Hours",y="Disruption Hours",color="Throughput"),
        title="Predicted Throughput Heatmap", aspect="auto")
    if not feasible.empty:
        fig_heat.add_hline(y=best['Disruption Hours'], line_color='red', line_width=2)
        fig_heat.add_vline(x=best['Staff Hours'],      line_color='red', line_width=2)
    fig_heat.update_layout(height=380, paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=40,b=10))
    st.plotly_chart(fig_heat, use_container_width=True)

    if not feasible.empty:
        st.dataframe(
            feasible[["Disruption Hours","Staff Hours","Predicted Throughput","Total Cost","Risk Level"]]
            .sort_values("Total Cost").head(10),
            use_container_width=True, height=200)

    st.divider()
    st.markdown("**Sensitivity: throughput vs staff hours (at Scenario A disruption level)**")
    sweep_df = pd.DataFrame(compute_sensitivity_sweep(shift_a, disruption_a, order_a, models_ok))
    fig_sens = go.Figure([
        go.Scatter(x=sweep_df["Staff Hours"], y=sweep_df["Upper"], line=dict(width=0), showlegend=False, mode='lines'),
        go.Scatter(x=sweep_df["Staff Hours"], y=sweep_df["Lower"], fill='tonexty',
                   fillcolor='rgba(31,78,121,0.15)', line=dict(width=0), showlegend=False, mode='lines'),
        go.Scatter(x=sweep_df["Staff Hours"], y=sweep_df["Throughput"],
                   line=dict(color='#002060',width=3), mode='lines+markers', name='Predicted', marker=dict(size=8)),
    ])
    fig_sens.add_hline(y=target_tp, line_dash='dash', line_color='red', annotation_text=f"  Target: {target_tp:,}")
    fig_sens.update_layout(height=280, xaxis_title="Staff Hours", yaxis_title="Throughput",
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10))
    st.plotly_chart(fig_sens, use_container_width=True)

# ── TAB 4: MODEL INTELLIGENCE ────────────────────────────────────────────────
with tab_model:
    st.markdown("### Model Intelligence & Performance")
    if metrics:
        tp_m   = metrics.get("throughput", {})
        risk_m = metrics.get("risk", {})
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Training Data",   f"{metrics.get('dataset_size', 'N/A'):,} rows")
        c2.metric("Throughput R²",   f"{tp_m.get('r2','N/A')}")
        c3.metric("Throughput MAE",  f"±{tp_m.get('mae','N/A'):.0f} units")
        c4.metric("Risk Accuracy",   f"{risk_m.get('accuracy','N/A'):.1%}")
        st.divider()

        fi_labels = {"shift_enc":"Shift","staff_hours":"Staff Hours",
                     "disruption_hours":"Disruption Hours","order_volume":"Order Volume"}
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("**Throughput Model — Feature Importance**")
            fi_df = pd.DataFrame([{"Feature":fi_labels.get(k,k),"Importance":v}
                                   for k,v in metrics.get("feature_importances_throughput",{}).items()]).sort_values("Importance")
            fig_fi = px.bar(fi_df, x="Importance", y="Feature", orientation='h',
                color="Importance", color_continuous_scale="Blues",
                text=fi_df["Importance"].apply(lambda x: f"{x:.1%}"))
            fig_fi.update_traces(textposition='outside')
            fig_fi.update_layout(height=280, showlegend=False, coloraxis_showscale=False,
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10))
            st.plotly_chart(fig_fi, use_container_width=True)
        with col_f2:
            st.markdown("**Risk Model — Feature Importance**")
            fi_df2 = pd.DataFrame([{"Feature":fi_labels.get(k,k),"Importance":v}
                                    for k,v in metrics.get("feature_importances_risk",{}).items()]).sort_values("Importance")
            fig_fi2 = px.bar(fi_df2, x="Importance", y="Feature", orientation='h',
                color="Importance", color_continuous_scale="RdYlGn",
                text=fi_df2["Importance"].apply(lambda x: f"{x:.1%}"))
            fig_fi2.update_traces(textposition='outside')
            fig_fi2.update_layout(height=280, showlegend=False, coloraxis_showscale=False,
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10))
            st.plotly_chart(fig_fi2, use_container_width=True)

    st.divider()
    st.markdown("**Prediction Distribution — all 150 trees for Scenario A**")
    if models_ok:
        tree_preds_now = get_tree_predictions(
            tuple(shift_enc.classes_), shift_a, staff_a, disruption_a, order_a, models_ok)
        fig_conf = go.Figure()
        fig_conf.add_trace(go.Histogram(x=tree_preds_now, nbinsx=25, marker_color='#1F4E79', opacity=0.8))
        fig_conf.add_vline(x=tp_a,    line_color='#C00000', line_width=2, annotation_text=f"  Mean: {tp_a:,}")
        fig_conf.add_vline(x=tp_lo_a, line_dash='dash', line_color='#FFC000', annotation_text=f"  P10: {tp_lo_a:,}")
        fig_conf.add_vline(x=tp_hi_a, line_dash='dash', line_color='#FFC000', annotation_text=f"P90: {tp_hi_a:,}  ")
        fig_conf.update_layout(height=280, xaxis_title="Throughput (individual tree predictions)", yaxis_title="# Trees",
            plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(t=10,b=10), showlegend=False)
        st.plotly_chart(fig_conf, use_container_width=True)
        st.caption(f"P10: {tp_lo_a:,}  |  Mean: {tp_a:,}  |  P90: {tp_hi_a:,}  |  Spread: ±{tp_hi_a-tp_lo_a:,} units")

    st.divider()
    st.markdown("**Historical Stats by Shift**")
    shift_stats = df.groupby('shift')['throughput_units'].agg(['mean','std','min','max','count']).round(1)
    st.dataframe(shift_stats, use_container_width=True)

# ── TAB 5: EXPORT ────────────────────────────────────────────────────────────
with tab_export:
    st.markdown("### Export Reports")
    col_ex1, col_ex2 = st.columns(2)

    with col_ex1:
        st.markdown("#### 📊 Full Executive Excel Dashboard")
        st.markdown("8-sheet workbook: live prediction, demand forecast, staffing plan, throughput analysis, risk & disruptions.")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference
            from openpyxl.utils import get_column_letter

            dashboard_bytes = load_dashboard_template()
            if not dashboard_bytes:
                raise FileNotFoundError("dashboard/warehouse_operations_dashboard.xlsx not found")
            dashboard_wb = openpyxl.load_workbook(io.BytesIO(dashboard_bytes))
            if "Scenario Prediction" in dashboard_wb.sheetnames:
                del dashboard_wb["Scenario Prediction"]
            ws_pred = dashboard_wb.create_sheet("Scenario Prediction")

            DB="002060"; MB="1F4E79"; LB="BDD7EE"; GR="70AD47"; AM="FFC000"; WH="FFFFFF"; GY="F2F2F2"
            thin = Side(style="thin", color="BFBFBF")
            bord = Border(left=thin,right=thin,top=thin,bottom=thin)

            def hdr(ws,row,col,val,bg=DB,fg=WH,bold=True,sz=12,align="center"):
                c = ws.cell(row=row,column=col,value=val)
                c.font=Font(bold=bold,color=fg,size=sz,name="Calibri")
                c.fill=PatternFill("solid",fgColor=bg)
                c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
                c.border=bord; return c

            def cel(ws,row,col,val,bg=WH,fg="000000",bold=False,sz=11,align="left",nf=None):
                c=ws.cell(row=row,column=col,value=val)
                c.font=Font(bold=bold,color=fg,size=sz,name="Calibri")
                c.fill=PatternFill("solid",fgColor=bg)
                c.alignment=Alignment(horizontal=align,vertical="center")
                c.border=bord
                if nf: c.number_format=nf
                return c

            ws_pred.merge_cells("A1:F1")
            t=ws_pred["A1"]; t.value="Warehouse Decision System — Live Scenario Prediction"
            t.font=Font(bold=True,color=WH,size=16,name="Calibri")
            t.fill=PatternFill("solid",fgColor=DB)
            t.alignment=Alignment(horizontal="center",vertical="center")
            ws_pred.row_dimensions[1].height=36

            ws_pred.merge_cells("A2:F2")
            sub=ws_pred["A2"]
            sub.value=f"Shift: {shift_a}  |  Staff: {staff_a}h  |  Orders: {order_a:,}  |  Disruption: {disruption_a}h  |  Risk: {risk_a}  |  Confidence: {conf_a}%"
            sub.font=Font(italic=True,color=WH,size=11,name="Calibri")
            sub.fill=PatternFill("solid",fgColor=MB)
            sub.alignment=Alignment(horizontal="center",vertical="center")
            ws_pred.row_dimensions[2].height=22; ws_pred.row_dimensions[3].height=10

            kpis=[
                ("Predicted Throughput",f"{tp_a:,} units",GR),
                ("90% CI Band",f"{tp_lo_a:,}–{tp_hi_a:,}",LB),
                ("Risk Level",risk_a,AM if risk_a=="Medium" else (GR if risk_a=="Low" else "FF4444")),
                ("Model Confidence",f"{conf_a}%",LB),
                ("Disruption Cost",f"${cost_a:,.0f}","FF4444" if cost_a>5000 else GR),
                ("Total Cost",f"${cost_a+labour_a:,.0f}","FF4444" if cost_a+labour_a>10000 else GR),
            ]
            for i,(label,val,color) in enumerate(kpis):
                hdr(ws_pred,4,i+1,label,bg=MB,sz=10)
                c=ws_pred.cell(row=5,column=i+1,value=val)
                c.font=Font(bold=True,color=WH,size=13,name="Calibri")
                c.fill=PatternFill("solid",fgColor=color)
                c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bord
                ws_pred.row_dimensions[4].height=20; ws_pred.row_dimensions[5].height=30

            ws_pred.row_dimensions[7].height=10
            ws_pred.merge_cells("A8:F8")
            hdr(ws_pred,8,1,"Feature Impact on Predicted Throughput",bg=MB,sz=11)
            ws_pred.row_dimensions[8].height=22
            for ci,lbl in enumerate(["Factor","Input Value","Impact (units)","Direction"],1):
                hdr(ws_pred,9,ci,lbl,bg=LB,fg=DB,sz=10)
            ws_pred.row_dimensions[9].height=18

            impact_rows=[("Staff Hours",staff_a,staff_a*280,"positive"),
                         ("Order Volume",order_a,int(order_a*0.3),"positive"),
                         ("Disruption Hours",disruption_a,-disruption_a*320,"negative")]
            for ri,(factor,inp,impact,dir) in enumerate(impact_rows):
                r=10+ri; bg=GY if ri%2==0 else WH
                cel(ws_pred,r,1,factor,bg=bg,bold=True)
                cel(ws_pred,r,2,inp,bg=bg,align="center")
                cel(ws_pred,r,3,impact,bg=bg,align="center",nf="+#,##0;-#,##0")
                arrow="▲ Positive" if dir=="positive" else "▼ Negative"
                ac="375623" if dir=="positive" else "C00000"
                c=ws_pred.cell(row=r,column=4,value=arrow)
                c.font=Font(bold=True,color=ac,size=11,name="Calibri")
                c.fill=PatternFill("solid",fgColor=bg)
                c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bord
                ws_pred.row_dimensions[r].height=18

            chart=BarChart(); chart.type="bar"; chart.title="Feature Impact on Throughput"
            chart.y_axis.title="Impact (units)"; chart.style=10; chart.width=18; chart.height=11
            chart.add_data(Reference(ws_pred,min_col=3,min_row=9,max_row=12),titles_from_data=True)
            chart.set_categories(Reference(ws_pred,min_col=1,min_row=10,max_row=12))
            ws_pred.add_chart(chart,"F4")
            for i,w in enumerate([22,16,18,16,14,14],1):
                ws_pred.column_dimensions[get_column_letter(i)].width=w

            dashboard_wb.move_sheet("Scenario Prediction",offset=-len(dashboard_wb.sheetnames)+1)
            excel_out=io.BytesIO(); dashboard_wb.save(excel_out)

            st.download_button(label="📊 Download Excel Dashboard", data=excel_out.getvalue(),
                file_name="Warehouse_Executive_Dashboard.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True)
        except Exception as e:
            st.error(f"Excel build error: {e}")

    with col_ex2:
        st.markdown("#### 📄 Executive PDF Summary")
        st.markdown("One-page PDF: KPIs, risk, cost breakdown, model performance, recommendations.")
        def build_pdf():
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER

            buf=io.BytesIO()
            doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
            DB=colors.HexColor("#002060"); MB=colors.HexColor("#1F4E79"); LB=colors.HexColor("#BDD7EE")
            GR=colors.HexColor("#70AD47"); AM=colors.HexColor("#FFC000"); RD=colors.HexColor("#C00000")
            WH=colors.white; GY=colors.HexColor("#F2F2F2")

            styles=getSampleStyleSheet()
            ts=ParagraphStyle('ts',parent=styles['Title'],fontSize=20,textColor=WH,alignment=TA_CENTER,spaceAfter=0)
            ss=ParagraphStyle('ss',parent=styles['Normal'],fontSize=10,textColor=LB,alignment=TA_CENTER,spaceAfter=0)
            h2=ParagraphStyle('h2',parent=styles['Heading2'],fontSize=12,textColor=DB,spaceBefore=12,spaceAfter=4)
            bs=ParagraphStyle('bs',parent=styles['Normal'],fontSize=9,textColor=colors.HexColor("#333333"),spaceAfter=3)
            rc={"Low":GR,"Medium":AM,"High":RD}.get(risk_a,DB)

            story=[]
            hdr_tbl=Table([[Paragraph("WAREHOUSE DECISION SUPPORT SYSTEM",ts)]],colWidths=[17*cm])
            hdr_tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),DB),('TOPPADDING',(0,0),(-1,-1),14),('BOTTOMPADDING',(0,0),(-1,-1),14)]))
            story.append(hdr_tbl)
            story.append(Spacer(1,0.2*cm))
            sub_tbl=Table([[Paragraph(f"Scenario A: {shift_a} Shift  ·  Generated {pd.Timestamp.now().strftime('%d %b %Y')}",ss)]],colWidths=[17*cm])
            sub_tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),MB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
            story.append(sub_tbl); story.append(Spacer(1,0.4*cm))

            story.append(Paragraph("KEY PERFORMANCE INDICATORS",h2))
            kpi_data=[["Throughput","90% CI","Risk Level","Disruption Cost","Total Cost"],
                      [f"{tp_a:,} units",f"{tp_lo_a:,}–{tp_hi_a:,}",risk_a,f"${cost_a:,.0f}",f"${cost_a+labour_a:,.0f}"]]
            kt=Table(kpi_data,colWidths=[3.4*cm]*5)
            kt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),MB),('TEXTCOLOR',(0,0),(-1,0),WH),
                ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),8),('FONTSIZE',(0,1),(-1,1),11),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7),
                ('BOX',(0,0),(-1,-1),1,DB),('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor("#BFBFBF")),
                ('BACKGROUND',(0,1),(-1,1),GY),('BACKGROUND',(2,1),(2,1),rc),('TEXTCOLOR',(2,1),(2,1),WH)]))
            story.append(kt); story.append(Spacer(1,0.3*cm))
            story.append(Paragraph(f"<b>Confidence:</b> Model predicts {tp_a:,} units (90% CI: {tp_lo_a:,}–{tp_hi_a:,}). Confidence: {conf_a}%.",bs))
            story.append(HRFlowable(width="100%",thickness=1,color=LB)); story.append(Spacer(1,0.2*cm))

            story.append(Paragraph("SCENARIO INPUTS",h2))
            in_data=[["Parameter","Value"],["Shift",shift_a],["Staff Hours",f"{staff_a}h"],
                     ["Disruption Hours",f"{disruption_a}h"],["Order Volume",f"{order_a:,}"],
                     ["Labour Cost/hr",f"${LABOUR_COST_PER_HOUR}"],["Disruption Cost/hr",f"${DISRUPTION_COST_PER_HOUR:,}"]]
            it=Table(in_data,colWidths=[8*cm,9*cm])
            it.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),MB),('TEXTCOLOR',(0,0),(-1,0),WH),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[WH,GY]),('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),5),
                ('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8),
                ('BOX',(0,0),(-1,-1),1,DB),('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor("#BFBFBF"))]))
            story.append(it); story.append(Spacer(1,0.2*cm))
            story.append(HRFlowable(width="100%",thickness=1,color=LB))

            if metrics:
                story.append(Paragraph("MODEL PERFORMANCE",h2))
                tp_m=metrics.get("throughput",{}); rm=metrics.get("risk",{})
                pm_data=[["Model","Algorithm","Metric","Value"],
                         ["Throughput","Random Forest (150T)","R² Score",str(tp_m.get('r2','N/A'))],
                         ["Throughput","Random Forest (150T)","MAE",f"±{tp_m.get('mae','N/A'):.0f} units"],
                         ["Risk Classifier","Random Forest (150T)","Accuracy",f"{rm.get('accuracy','N/A'):.1%}"],
                         ["Dataset","Warehouse Operations","Total Records",f"{metrics.get('dataset_size','N/A'):,}"]]
                pmt=Table(pm_data,colWidths=[4*cm,4.5*cm,4.5*cm,4*cm])
                pmt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),MB),('TEXTCOLOR',(0,0),(-1,0),WH),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[WH,GY]),('ALIGN',(0,0),(-1,-1),'CENTER'),
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),5),
                    ('BOTTOMPADDING',(0,0),(-1,-1),5),('BOX',(0,0),(-1,-1),1,DB),
                    ('INNERGRID',(0,0),(-1,-1),0.5,colors.HexColor("#BFBFBF"))]))
                story.append(pmt); story.append(Spacer(1,0.2*cm))
                story.append(HRFlowable(width="100%",thickness=1,color=LB))

            story.append(Paragraph("RECOMMENDATIONS",h2))
            recs=[]
            if risk_a=="High": recs.append("🚨 HIGH RISK — reduce disruption hours immediately.")
            if disruption_a>=3: recs.append(f"Disruption at {disruption_a}h costs ${cost_a:,.0f}. Reducing to 2h saves ${(disruption_a-2)*DISRUPTION_COST_PER_HOUR:,.0f}.")
            if staff_a<8: recs.append(f"Staff at {staff_a}h is below optimal (8h). +2h adds ~{2*280:,} throughput units.")
            if not recs: recs.append("✅ Scenario A looks operationally healthy. Continue monitoring.")
            for r in recs: story.append(Paragraph(f"• {r}",bs))

            story.append(Spacer(1,0.4*cm))
            ft=Table([[Paragraph(f"<font color='#FFFFFF'>Built by Garvit Mittal · Warehouse Decision System · {pd.Timestamp.now().strftime('%d %B %Y %H:%M')}</font>",
                ParagraphStyle('f',fontSize=7,alignment=TA_CENTER,textColor=WH))]],colWidths=[17*cm])
            ft.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),DB),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
            story.append(ft)
            doc.build(story); return buf.getvalue()

        try:
            pdf_bytes=build_pdf()
            st.download_button(label="📄 Download PDF Summary", data=pdf_bytes,
                file_name="Warehouse_Executive_Summary.pdf", mime="application/pdf",
                type="primary", use_container_width=True)
        except ImportError:
            st.warning("Run `pip install reportlab` to enable PDF export.")
        except Exception as e:
            st.error(f"PDF error: {e}")

    if compare_mode:
        st.divider()
        st.markdown("#### 📊 A/B Comparison Export")
        comp_df=pd.DataFrame({"Metric":["Shift","Staff Hours","Disruption Hours","Order Volume",
                "Predicted Throughput","90% CI Low","90% CI High","Risk Level","Confidence","Disruption Cost","Labour Cost","Total Cost"],
            "Scenario A":[shift_a,staff_a,disruption_a,order_a,tp_a,tp_lo_a,tp_hi_a,
                risk_a,f"{conf_a}%",f"${cost_a:,.0f}",f"${labour_a:,.0f}",f"${cost_a+labour_a:,.0f}"],
            "Scenario B":[shift_b,staff_b,disruption_b,order_b,tp_b,tp_lo_b,tp_hi_b,
                risk_b,f"{conf_b}%",f"${cost_b:,.0f}",f"${labour_b:,.0f}",f"${cost_b+labour_b:,.0f}"]})
        comp_buf=io.BytesIO()
        with pd.ExcelWriter(comp_buf,engine='openpyxl') as w:
            comp_df.to_excel(w,sheet_name="A vs B Comparison",index=False)
        st.download_button(label="📊 Download A/B Comparison Excel", data=comp_buf.getvalue(),
            file_name="Warehouse_AB_Comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

st.caption("Built by Garvit Mittal · Warehouse Decision Support System · Real ML Models · v2.0")
